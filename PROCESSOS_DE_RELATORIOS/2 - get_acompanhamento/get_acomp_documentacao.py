import threading
import schedule
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
import time
import tempfile
import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from cred import OK_ENTREGA_URL, OK_ENTREGA_USUARIO, OK_ENTREGA_SENHA


def esperar_download_concluir(diretorio, timeout=60):
    for _ in range(timeout):
        arquivos = [f for f in os.listdir(diretorio) if f.endswith(".xlsx")]
        if arquivos:
            return arquivos[0]
        time.sleep(1)
    return None


def tratar_e_formatar_arquivo(arquivo_baixado, arquivo_destino):
    tempo_finalizado = [False]

    def mostrar_tempo():
        inicio = time.time()
        while not tempo_finalizado[0]:
            print(f"\r⏱ 😄 Processando... {int(time.time() - inicio)}s", end="", flush=True)
            time.sleep(1)

    thread = threading.Thread(target=mostrar_tempo)
    thread.start()

    df = pd.read_excel(arquivo_baixado, dtype=str)

    colunas_texto = [
        "CNPJ Embarcador",
        "Filial embarcador",
        "CNPJ Transportadora",
        "CNPJ Destinatário",
        "Chave Danfe",
        "Codigo IBGE Municipio"
    ]
    for col in colunas_texto:
        if col in df.columns:
            df[col] = df[col].astype(str).fillna("")

    colunas_data = [
        "Emissão",
        "Despacho",
        "Previsão entrega",
        "Data da Entrega",
        "Data canhoto"
    ]
    for col in colunas_data:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    with pd.ExcelWriter(arquivo_destino, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, index=False, sheet_name="Dados")

    wb = load_workbook(arquivo_destino)
    ws = wb.active

    fill = PatternFill(start_color="30CA70", end_color="30CA70", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(border_style="thin", color="D9D9D9"),
        right=Side(border_style="thin", color="D9D9D9"),
        top=Side(border_style="thin", color="D9D9D9"),
        bottom=Side(border_style="thin", color="D9D9D9")
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    wb.save(arquivo_destino)
    wb.close()

    tempo_finalizado[0] = True
    thread.join()
    print("\n✅ Formatação concluída.")
    print("Arquivo salvo em:")
    print(arquivo_destino)


def segundo_acesso_():
    inicio = time.time()
    temp_dir = tempfile.mkdtemp()

    options = Options()
    for arg in [
        "--start_maximized",
        "--disable-extensions",
        "--allow-running-insecure-content",
        f"--unsafely-treat-insecure-origin-as-secure={OK_ENTREGA_URL}",
        "--disable-gpu",
        "--no-sandbox",
        "--disable-dev-shm-usage"
    ]: options.add_argument(arg)

    options.add_experimental_option("prefs", {
        "download.default_directory": temp_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False,
    })

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    try:
        driver.get(OK_ENTREGA_URL)
        driver.maximize_window()

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "camp_identificacao"))
        ).send_keys(OK_ENTREGA_USUARIO)

        driver.find_element(By.XPATH, '//input[@placeholder="Informe sua senha ..."]').send_keys(OK_ENTREGA_SENHA)
        driver.find_element(By.CLASS_NAME, "loginButton").click()

        print("Entrando no sistema... Aguarde.")
        time.sleep(5)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "hamburger"))
        ).click()

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//a[span[contains(text(), "Consulta")]]'))
        ).click()
        time.sleep(1)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//a[text()="Exportação Excel"]'))
        ).click()
        time.sleep(5)

        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "icon_close"))
        ).click()

        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element_located((By.ID, "alertModalAviso"))
        )

        max_tentativas = 3
        tentativas = 0
        baixado = False

        while tentativas < max_tentativas and not baixado:
            try:
                tabela = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, "listaSolicitacoes"))
                )

                primeira_linha = tabela.find_element(By.CSS_SELECTOR, "tbody tr:first-child")
                status = primeira_linha.find_element(By.CSS_SELECTOR, "td:nth-child(2) > div > span").text.strip()
                print(f"Tentativa {tentativas + 1} - Status:", status)

                if status.lower() == "executado":
                    link_element = WebDriverWait(primeira_linha, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "td:nth-child(7) a"))
                    )
                    link_element.click()
                    print("Cliquei para baixar o arquivo.")

                    nome_arquivo = esperar_download_concluir(temp_dir, timeout=60)
                    if nome_arquivo:
                        caminho_baixado = os.path.join(temp_dir, nome_arquivo)
                        print("Arquivo baixado:", caminho_baixado)

                        tratar_e_formatar_arquivo(
                            caminho_baixado,
                            r"\\server\JTDTRANSPORTES2\ANALITCS\ACOMPANHAMENTO DE CARGAS\ok entregas.xlsx"
                        )
                        print("Arquivo tratado e salvo no destino.")
                        baixado = True
                    else:
                        print("O download não foi concluído dentro do tempo limite.")
                else:
                    tentativas += 1
                    if tentativas < max_tentativas:
                        print("Ainda não está 'Executado'. Aguardando 10 minutos para tentar novamente...")
                        time.sleep(600)
                        driver.refresh()
                        time.sleep(5)

                        try:
                            ok_btn = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="Ok"]'))
                            )
                            ok_btn.click()
                            WebDriverWait(driver, 10).until(
                                EC.invisibility_of_element_located((By.ID, "alertModalAviso"))
                            )
                        except:
                            pass
                        driver.execute_script("window.scrollBy(0, 800)")
                    else:
                        print("Número máximo de tentativas atingido. Encerrando.")
            except Exception as e:
                print("Erro geral no laço:", e)
                tentativas += 1
                time.sleep(5)

    except Exception as e:
        print(f"Erro durante a execução: {e}")
    finally:
        tempo_total = int(time.time() - inicio)
        print(f"\n⏳ Processo finalizado em {tempo_total}s.")
        time.sleep(5)
        driver.quit()

segundo_acesso_()

schedule.every().day.at("07:40").do(segundo_acesso_)
schedule.every().day.at("09:40").do(segundo_acesso_)
schedule.every().day.at("13:40").do(segundo_acesso_)
schedule.every().day.at("17:10").do(segundo_acesso_)

ultimo_log = 0

while True:
    schedule.run_pending()

    agora = time.time()
    if agora - ultimo_log >= 120:
        print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] \U0001F4D1 Aguardando CÓPIA")
        ultimo_log = agora

    time.sleep(1)
