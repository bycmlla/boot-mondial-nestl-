import time
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from google.oauth2.service_account import Credentials
import gspread
import os
import schedule


def access_api():
    print(os.path.isfile("credenciais.json"))
    print(os.path.getsize("credenciais.json"))
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    credentials = Credentials.from_service_account_file("credenciais.json", scopes=scopes)
    client = gspread.authorize(credentials)

    spreadsheet = client.open("MONITORAMENTO JTD")

    hoje = datetime.today()
    nome_aba = hoje.strftime("%d/%m/%Y")
    sheet = spreadsheet.worksheet(nome_aba)

    data = sheet.get_all_records()
    df = pd.DataFrame(data)

    ontem = hoje - timedelta(days=1)
    data_ontem = ontem.strftime("%d/%m/%Y")

    df_filtrado = df[df['Data_expedicao'] == data_ontem]

    if df_filtrado.empty:
        print("Nenhum dado encontrado para ontem.")
    else:
        caminho_arquivo_destino = r"\\server\JTDTRANSPORTES2\ANALITCS\ACOMPANHAMENTO DE CARGAS\CARGAS DIRETA - MONDIAL.xlsx"
        sheet_name = 'Cargas'

        book = load_workbook(caminho_arquivo_destino)
        sheet_destino = book[sheet_name]

        ultima_linha = sheet_destino.max_row

        for index, row in df_filtrado.iterrows():
            sheet_destino.append(list(row))

        border = Border(left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))

        num_linhas_adicionadas = len(df_filtrado)
        primeira_linha_nova = ultima_linha + 1
        ultima_linha_nova = primeira_linha_nova + num_linhas_adicionadas - 1
        num_colunas = len(df_filtrado.columns)

        for row in sheet_destino.iter_rows(min_row=primeira_linha_nova, max_row=ultima_linha_nova, min_col=1,
                                           max_col=num_colunas):
            for cell in row:
                cell.border = border

        book.save(caminho_arquivo_destino)

        print(f"{num_linhas_adicionadas} linha(s) adicionada(s) na planilha de destino.")

# access_api()
schedule.every().day.at("10:00").do(access_api)


ultimo_log: int = 0

while True:
    schedule.run_pending()

    agora = time.time()
    if agora - ultimo_log >= 120:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] \U0001F310 Aguardando GOOGLE SHEETS")
        ultimo_log = agora

    time.sleep(1)
