import logging
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def format_excel_sheet(sheet, df_columns):
    logging.info(f"Formatando planilha (sem bordas/cores): {sheet.title}")

    col_indices = {name.lower(): i + 1 for i, name in enumerate(df_columns)}
    col_data_idx = col_indices.get("tcon_data_emissao")
    col_valor_idx = col_indices.get("tcon_valor_liquido")
    col_soma_valor_idx = col_indices.get("soma de valor líquido")
    col_mes_idx = col_indices.get("mes")

    logging.info(
        f"Índices para formatação ({sheet.title}): Data={col_data_idx}, Valor={col_valor_idx}, Soma Pivot={col_soma_valor_idx}, Mes Pivot={col_mes_idx}")

    min_data_row = 2
    col_mes_idx_sheet = None
    col_soma_valor_idx_sheet = None

    if sheet.title == 'ResumoMensal':
        col_mes_idx_sheet = 1
        col_soma_valor_idx_sheet = 2
    elif sheet.title == 'Dados':
        col_mes_idx_sheet = col_mes_idx
        col_soma_valor_idx_sheet = None

    for row in sheet.iter_rows(min_row=min_data_row, max_row=sheet.max_row):
        for cell in row:
            if sheet.title == 'Dados' and col_data_idx and cell.column == col_data_idx and cell.value is not None:
                try:
                    cell.number_format = 'DD/MM/YYYY'
                except ValueError:
                    logging.warning(
                        f"Não foi possível aplicar formato de data à célula {cell.coordinate} com valor {cell.value}")

            if sheet.title == 'Dados' and col_valor_idx and cell.column == col_valor_idx and cell.value is not None:
                try:
                    cell.number_format = 'R$ #,##0.00'
                except ValueError:
                    logging.warning(
                        f"Não foi possível aplicar formato de moeda à célula {cell.coordinate} com valor {cell.value}")

            if sheet.title == 'ResumoMensal' and col_soma_valor_idx_sheet and cell.column == col_soma_valor_idx_sheet and cell.value is not None:
                try:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = 'R$ #,##0.00'
                except ValueError:
                    logging.warning(
                        f"Não foi possível aplicar formato de moeda à célula {cell.coordinate} (pivot) com valor {cell.value}")

            if sheet.title == 'Dados' and col_mes_idx_sheet and cell.column == col_mes_idx_sheet and cell.value is not None:
                try:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0'
                except ValueError:
                    pass

    logging.info(f"Ajustando largura das colunas para {sheet.title}...")
    for i, column_cells in enumerate(sheet.columns):
        try:
            max_length = 0
            column_letter = get_column_letter(i + 1)
            current_col_index = i + 1

            is_currency_col = (sheet.title == 'Dados' and current_col_index == col_valor_idx) or \
                              (sheet.title == 'ResumoMensal' and current_col_index == col_soma_valor_idx_sheet)
            is_date_col = sheet.title == 'Dados' and current_col_index == col_data_idx
            is_month_name_col = sheet.title == 'ResumoMensal' and current_col_index == col_mes_idx_sheet

            header_rows = 1
            if sheet.title == 'ResumoMensal':
                if sheet.cell(row=2, column=1).value is None and isinstance(sheet.cell(row=1, column=1).value,
                                                                            str):
                    header_rows = 1
            min_data_row_for_width = header_rows + 1

            for row_idx in range(1, min_data_row_for_width):
                cell_value = sheet.cell(row=row_idx, column=current_col_index).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            for cell in column_cells[min_data_row_for_width - 1:]:
                if cell.value is not None:
                    if is_date_col:
                        cell_len = 10
                    elif is_month_name_col:
                        cell_len = len(str(cell.value))
                    elif is_currency_col and isinstance(cell.value, (int, float)):
                        try:
                            formatted_value = f"R$ {cell.value:,.2f}".replace(",", "v").replace(".", ",").replace("v",
                                                                                                                  ".")
                            cell_len = len(formatted_value) + 1
                        except:
                            cell_len = len(str(cell.value))
                    else:
                        cell_len = len(str(cell.value))

                    max_length = max(max_length, cell_len)

            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            logging.warning(f"Erro ao ajustar largura da coluna {i + 1} ({get_column_letter(i + 1)}): {e}")

    logging.info(f"Formatação essencial da planilha {sheet.title} concluída.")
