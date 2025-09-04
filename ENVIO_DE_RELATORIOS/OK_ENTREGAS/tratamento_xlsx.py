import os
import shutil
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import time
import threading

def tratar_e_formatar_arquivo(temp_dir, nome_arquivo):
    tempo_finalizado = [False]

    def mostrar_tempo():
        inicio = time.time()
        while not tempo_finalizado[0]:
            print(f"\r⏱ 😄 Processando... {int(time.time() - inicio)}s", end="", flush=True)
            time.sleep(1)

    thread = threading.Thread(target=mostrar_tempo)
    thread.start()

    try:
        caminho_original = os.path.join(temp_dir, nome_arquivo)
        novo_nome = "ACOMPANHAMENTO OK ENTREGAS.xlsx"
        caminho_final_temp = os.path.join(temp_dir, novo_nome)
        caminho_final_downloads = os.path.join(os.path.expanduser("~/Downloads"), novo_nome)

        df = pd.read_excel(caminho_original, dtype=str)
        print("\nComeçando tratamento! 😄")

        if "Filial embarcador" in df.columns:
            df["Filial embarcador"] = df["Filial embarcador"].str.zfill(4)

        campos_data = ["Emissão", "Despacho", "Previsão entrega", "Previsão de entrega Original"]
        for campo in campos_data:
            if campo in df.columns:
                df[campo] = pd.to_datetime(df[campo], errors="coerce").dt.strftime("%d/%m/%Y")

        campos_cnpj = ["CNPJ Transportadora", "CNPJ Destinatário", "CNPJ Embarcador"]
        for campo in campos_cnpj:
            if campo in df.columns:
                df[campo] = df[campo].astype(str).str.zfill(14)

        df_base = df.copy()

        df_filtros = df[df["Canhoto"] != "REGISTRADO"]

        embarcadores_validos = ["mk", "casas bahia", "americanas", "magazine luiza", "claudino", "n.claudino"]
        if "Embarcador" in df_filtros.columns:
            df_filtros = df_filtros[df_filtros["Embarcador"].str.lower().str.contains(
                "|".join(embarcadores_validos), na=False
            )]

        with pd.ExcelWriter(caminho_final_temp, engine="openpyxl") as writer:
            df_base.to_excel(writer, index=False, sheet_name="Base")
            df_filtros.to_excel(writer, index=False, sheet_name="Filtros")

        wb = load_workbook(caminho_final_temp)

        ws_filtros = wb["Filtros"]
        ws_filtros.title = "Acompanhamento"
        ws_acompanhamento = wb["Acompanhamento"]

        wb._sheets.remove(ws_acompanhamento)
        wb._sheets.insert(0, ws_acompanhamento)

        verde = "30CA70"
        branco = "FFFFFF"
        cinza = "D9D9D9"
        vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        azul = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

        header_fill = PatternFill(start_color=verde, end_color=verde, fill_type="solid")
        header_font = Font(color=branco, bold=True, name="Arial", size=10)
        body_font = Font(name="Arial", size=10)
        border = Border(
            left=Side(border_style="thin", color=cinza),
            right=Side(border_style="thin", color=cinza),
            top=Side(border_style="thin", color=cinza),
            bottom=Side(border_style="thin", color=cinza)
        )

        def formatar_planilha(ws):
            hoje = datetime.now().date()
            headers = [cell.value for cell in ws[1]]
            previsao_col_index = headers.index("Previsão entrega") + 1 if "Previsão entrega" in headers else None

            for col_num, cell in enumerate(ws[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 18

            for row in ws.iter_rows(min_row=2):
                for col_index, cell in enumerate(row, 1):
                    cell.font = body_font
                    cell.border = border
                    if previsao_col_index and col_index == previsao_col_index:
                        try:
                            data = datetime.strptime(cell.value, "%d/%m/%Y").date()
                            dias = (data - hoje).days
                            if dias <= 5:
                                cell.fill = vermelho
                            elif dias <= 10:
                                cell.fill = amarelo
                            elif dias <= 15:
                                cell.fill = azul
                        except:
                            pass

        formatar_planilha(ws_acompanhamento)

        wb.save(caminho_final_temp)
        shutil.copy(caminho_final_temp, caminho_final_downloads)

        print("\n✅ Formatação da aba 'Acompanhamento' concluída.")
        print("✅ Arquivo tratado salvo em:")
        print("-", caminho_final_temp)
        print("-", caminho_final_downloads)

        return caminho_final_temp

    finally:
        tempo_finalizado[0] = True
        thread.join()
