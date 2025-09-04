import logging
import os
import glob
import zipfile
import pandas as pd
import shutil
from datetime import datetime
from openpyxl import load_workbook
import warnings

warnings.simplefilter(action='ignore', category=pd.errors.SettingWithCopyWarning)

logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')

def descompactar_csv(download_dir):
    try:
        zip_files = glob.glob(os.path.join(download_dir, "*.zip"))
        if not zip_files:
            return None

        zip_file_path = zip_files[0]
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            zip_ref.extractall(download_dir)

        csv_files = glob.glob(os.path.join(download_dir, "*.csv"))
        if not csv_files:
            return None

        csv_file_path = csv_files[0]

        try:
            df = pd.read_csv(csv_file_path, sep=';', encoding='cp1252', on_bad_lines='warn')
        except UnicodeDecodeError:
            df = pd.read_csv(csv_file_path, sep=';', encoding='utf-8', on_bad_lines='warn')

        if "CLI_RAZAO_SOCIAL" not in df.columns:
            logging.error(f"Coluna 'CLI_RAZAO_SOCIAL' não encontrada no arquivo CSV.")
            return None

        df["CLI_RAZAO_SOCIAL"] = df["CLI_RAZAO_SOCIAL"].astype(str).str.strip()
        df_filtrado = df[df["CLI_RAZAO_SOCIAL"].str.lower().str.contains("nestle", na=False)]

        df_filtrado["TCON_DATA_EMISSAO"] = pd.to_datetime(df_filtrado["TCON_DATA_EMISSAO"], dayfirst=True, errors='coerce').dt.date
        df_filtrado["TCON_VALOR_LIQUIDO"] = df_filtrado["TCON_VALOR_LIQUIDO"].astype(str).str.replace('.', '', regex=False).str.replace(',', '.')
        df_filtrado["TCON_VALOR_LIQUIDO"] = pd.to_numeric(df_filtrado["TCON_VALOR_LIQUIDO"], errors='coerce').fillna(0)

        resumo = df_filtrado.groupby("TCON_DATA_EMISSAO")["TCON_VALOR_LIQUIDO"].sum().reset_index()
        resumo.rename(columns={
            "TCON_DATA_EMISSAO": "DATA",
            "TCON_VALOR_LIQUIDO": "VALOR LIQUIDO"
        }, inplace=True)

        nome_aba_base = "Base"
        nome_aba_dinamica = "Dinâmica"

        data_hoje_str = datetime.now().strftime("%d-%m-%Y")
        nome_arquivo_excel = os.path.join(download_dir, f"NESTLÉ {data_hoje_str}.xlsx")

        with pd.ExcelWriter(nome_arquivo_excel, engine="openpyxl") as writer:
            df_filtrado.to_excel(writer, index=False, sheet_name=nome_aba_base)
            resumo.to_excel(writer, index=False, sheet_name=nome_aba_dinamica)

        wb = load_workbook(nome_arquivo_excel)
        ws = wb[nome_aba_dinamica]

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'

        wb.save(nome_arquivo_excel)
        wb.close()

        nome_saida_csv = os.path.join(download_dir, "filtrado_nestle.csv")
        df_filtrado.to_csv(nome_saida_csv, sep=";", index=False, encoding="utf-8-sig")

        return nome_arquivo_excel

    except zipfile.BadZipFile:
        logging.error(f"Erro ao descompactar: O arquivo {zip_file_path} não é um ZIP válido ou está corrompido.")
        return None
    except Exception as e:
        logging.error(f"Erro inesperado ao descompactar e tratar CSV: {e}", exc_info=True)
        return None
